"""隨機抽測核心邏輯：開真瀏覽器打真實聊天頁面、發問、截圖、記錄結果。
持續執行，每小時隨機抽測 2~4 次，直到按 ESC（有終端機時）或收到停止訊號為止。"""
import json
import os
import random
import select
import sys
import time
import traceback
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright, Route, Request

from question_bank import fetch_question_bank
from tenants import load_tenants_by_category

TAIWAN_TZ = timezone(timedelta(hours=8))


def taiwan_now() -> datetime:
    return datetime.now(TAIWAN_TZ)

TARGET_API_URL = os.environ.get("TARGET_API_URL", "http://localhost:8080").rstrip("/")
INTERNAL_API_URL = os.environ.get("INTERNAL_API_URL", "http://nginx").rstrip("/")
NAV_BASE_URL = os.environ.get("QA_NAV_BASE_URL", INTERNAL_API_URL)

QA_TENANT_CATEGORY = os.environ.get("QA_TENANT_CATEGORY", "car")
QA_TESTS_PER_HOUR_MIN = int(os.environ.get("QA_TESTS_PER_HOUR_MIN", "2"))
QA_TESTS_PER_HOUR_MAX = int(os.environ.get("QA_TESTS_PER_HOUR_MAX", "4"))
QA_TENANTS_PER_TEST_MIN = int(os.environ.get("QA_TENANTS_PER_TEST_MIN", "3"))
QA_TENANTS_PER_TEST_MAX = int(os.environ.get("QA_TENANTS_PER_TEST_MAX", "7"))
QA_DELAY_BETWEEN_TESTS_MIN = float(os.environ.get("QA_DELAY_BETWEEN_TESTS_MIN", "8"))
QA_DELAY_BETWEEN_TESTS_MAX = float(os.environ.get("QA_DELAY_BETWEEN_TESTS_MAX", "20"))
QA_TARGET_LANGS = [l.strip() for l in os.environ.get("QA_TARGET_LANGS", "zh-tw,en,ja,ko,vi,id,th").split(",") if l.strip()]
QA_MOCK_GEMINI = os.environ.get("QA_MOCK_GEMINI", "false").lower() == "true"

OUTPUT_ROOT = Path(os.environ.get("QA_OUTPUT_DIR", "/app/data/qa-verification"))

EXCEL_HEADERS = [
    "timestamp", "tenant_id", "question_id", "lang", "question",
    "expected_answer", "actual_response", "success", "screenshot_path", "duration_ms", "error",
]

# src/locales/translations.json 裡各語言的通用錯誤訊息，用來判定這輪問答是否失敗
ERROR_MESSAGES = {
    "zh-tw": "抱歉，AI 模型暫時無法回應，請稍後再試一次~~~",
    "en": "Sorry, I encountered an issue. Please try again later.",
    "ja": "申し訳ございません。問題が発生しました。後ほど再度お試しください。",
    "ko": "죄송합니다. 문제가 발생했습니다. 나중에 다시 시도해 주세요.",
    "vi": "Xin lỗi, tôi gặp sự cố. Vui lòng thử lại sau.",
    "id": "Maaf, saya mengalami masalah. Silakan coba lagi nanti.",
    "th": "ขออภัยค่ะ เกิดปัญหาขึ้น กรุณาลองใหม่อีกครั้งในภายหลัง",
}


def make_route_handler(lang: str, mock_answer: str = None):
    """把寫死在前端建置的 TARGET_API_URL 改寫成 container 內部可連到的 INTERNAL_API_URL，
    並在送出 /api/chat 時把 lang 換成這次隨機抽到的語言，藉此測試多語回覆而不必額外翻譯題目。
    若有帶 mock_answer，/api/chat 直接在瀏覽器端攔截回傳假回答，完全不會打到後端/Gemini，
    純粹用來測試「發問 -> 畫面渲染 -> 截圖 -> 記錄」這條自動化流程本身，不受配額/限流影響。"""

    def handle(route: Route, request: Request):
        url = request.url
        new_url = url
        if TARGET_API_URL != INTERNAL_API_URL and url.startswith(TARGET_API_URL):
            new_url = INTERNAL_API_URL + url[len(TARGET_API_URL):]

        if request.method == "POST" and url.endswith("/api/chat"):
            if mock_answer is not None:
                route.fulfill(
                    status=200,
                    content_type="application/json",
                    body=json.dumps({"response": mock_answer, "references": []}, ensure_ascii=False),
                )
                return
            try:
                data = json.loads(request.post_data or "{}")
                data["lang"] = None if lang == "zh-tw" else lang
                route.continue_(url=new_url, post_data=json.dumps(data))
                return
            except Exception:
                pass

        route.continue_(url=new_url)

    return handle


def run_single_test(playwright, tenant_id: str, lang: str, q: dict, mock_answer: str = None) -> dict:
    browser = playwright.chromium.launch(headless=True)
    context = browser.new_context(viewport={"width": 480, "height": 800})
    page = context.new_page()
    # 保留瀏覽器端的 JS 錯誤 log，方便之後排查「畫面卡在錯誤訊息但後端其實沒出錯」這類問題
    page.on("console", lambda msg: print(f"[browser-console] {msg.text}", flush=True) if msg.type == "error" else None)
    page.on("pageerror", lambda exc: print(f"[browser-pageerror] {exc}", flush=True))
    page.route("**/*", make_route_handler(lang, mock_answer))

    start = time.time()
    result = {
        "question_id": q["id"],
        "tenant_id": tenant_id,
        "lang": lang,
        "question": q["question"],
        "expected_answer": q["expected_answer"],
        "timestamp": taiwan_now().isoformat(),
    }

    try:
        page.goto(f"{NAV_BASE_URL}/chat?tenant_id={tenant_id}", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_selector("[data-qa-message-list] [data-message-id]", timeout=20000)

        # 送出問題後，使用者訊息會「立刻」出現在 DOM，但那不是機器人回答——
        # 一定要明確等到一個新的 bot_/error_ id 出現，不能只看訊息總數有沒有增加
        initial_ids = page.eval_on_selector_all(
            "[data-qa-message-list] [data-message-id]", "els => els.map(el => el.getAttribute('data-message-id'))"
        )

        input_box = page.locator('input[type="text"]')
        input_box.wait_for(state="visible", timeout=20000)
        input_box.fill(q["question"])
        page.locator('button[type="submit"]').click()

        page.wait_for_function(
            """(initialIds) => {
                const ids = Array.from(document.querySelectorAll('[data-qa-message-list] [data-message-id]'))
                    .map(el => el.getAttribute('data-message-id'));
                return ids.some(id => (id.startsWith('bot_') || id.startsWith('error_')) && !initialIds.includes(id));
            }""",
            arg=initial_ids,
            timeout=45000,
        )

        bubbles = page.locator("[data-qa-message-list] [data-message-id]")
        count = bubbles.count()
        user_el = bubbles.nth(count - 2)
        bot_el = bubbles.nth(count - 1)

        # 新訊息出現後，聊天室會用 smooth scroll 捲到底部，這個動畫還沒結束時就量測座標，
        # 會導致 bounding box 跟實際截圖當下的畫面位置對不上（截到捲動途中的錯誤區域，
        # 例如只截到最上方的歡迎訊息）。用 Playwright 自己的立即捲動先穩定位置，再等排版settle。
        bot_el.scroll_into_view_if_needed()
        page.wait_for_timeout(300)

        response_text = bot_el.inner_text().strip()

        user_box = user_el.bounding_box()
        bot_box = bot_el.bounding_box()
        clip = {
            "x": 0,
            "y": max(user_box["y"], 0),
            "width": page.viewport_size["width"],
            "height": (bot_box["y"] + bot_box["height"]) - user_box["y"],
        }

        now_tw = taiwan_now()
        date_str = now_tw.strftime("%Y%m%d")
        tenant_dir = OUTPUT_ROOT / date_str / tenant_id
        tenant_dir.mkdir(parents=True, exist_ok=True)
        # 檔名統一用「月日_時分秒_毫秒」（台灣時區），毫秒只是為了避免同一秒內撞名
        screenshot_path = tenant_dir / f"{q['id']}_{lang}_{now_tw.strftime('%m%d_%H%M%S_%f')[:-3]}.png"
        page.screenshot(path=str(screenshot_path), clip=clip)

        # inner_text() 會把訊息內容跟下方時間戳記一起抓進來，所以用「包含」而非完全相等比對
        is_error = any(err_msg in response_text for err_msg in ERROR_MESSAGES.values())
        result.update({
            "actual_response": response_text,
            "success": bool(response_text) and not is_error,
            "screenshot_path": str(screenshot_path),
            "duration_ms": int((time.time() - start) * 1000),
        })
    except Exception as e:
        traceback.print_exc()
        result.update({
            "actual_response": None,
            "success": False,
            "error": f"{type(e).__name__}: {e}",
            "screenshot_path": None,
            "duration_ms": int((time.time() - start) * 1000),
        })
    finally:
        context.close()
        browser.close()

    return result


def append_result_excel(result: dict):
    date_str = taiwan_now().strftime("%Y%m%d")
    day_dir = OUTPUT_ROOT / date_str
    day_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = day_dir / "results.xlsx"

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "results"

    # 檔案剛建立、或使用者自己先建了一個空白 Excel 檔佔位，都需要先補表頭。
    # 全新的 openpyxl 工作表即使沒寫過任何東西，max_row 也會回報 1，
    # 直接 append 表頭會落到第 2 列、留下一列空白，所以先清掉那個「假的」第 1 列
    if ws["A1"].value is None:
        ws.delete_rows(1, ws.max_row)
        ws.append(EXCEL_HEADERS)

    ws.append([result.get(h, "") for h in EXCEL_HEADERS])
    wb.save(xlsx_path)


def wait_with_esc_check(seconds: float, is_tty: bool) -> bool:
    """休息指定秒數；有互動式終端機時每 0.2 秒輪詢一次鍵盤，按 ESC 就提早回傳 True 要求結束"""
    if not is_tty:
        time.sleep(seconds)
        return False

    step = 0.2
    elapsed = 0.0
    while elapsed < seconds:
        ready, _, _ = select.select([sys.stdin], [], [], step)
        if ready:
            ch = sys.stdin.read(1)
            if ch == "\x1b":  # ESC
                return True
        elapsed += step
    return False


def run_forever():
    print(f"[qa-verifier] 開始連續隨機測試（每小時約 {QA_TESTS_PER_HOUR_MIN}~{QA_TESTS_PER_HOUR_MAX} 輪，"
          f"每輪隨機問 {QA_TENANTS_PER_TEST_MIN}~{QA_TENANTS_PER_TEST_MAX} 家租戶），"
          f"{'按 ESC 可隨時結束' if sys.stdin.isatty() else '背景執行中，用 docker stop 結束'}")
    if QA_MOCK_GEMINI:
        print("[qa-verifier] QA_MOCK_GEMINI=true，本次不會呼叫真正的後端/Gemini，只測試自動化流程本身")

    questions = fetch_question_bank()
    car_tenants = load_tenants_by_category(QA_TENANT_CATEGORY)
    print(f"[qa-verifier] 題庫共 {len(questions)} 題，{QA_TENANT_CATEGORY} 分類共 {len(car_tenants)} 個租戶")

    if not questions or not car_tenants:
        print("[qa-verifier] 題庫或租戶清單為空，無法執行")
        return

    is_tty = sys.stdin.isatty()

    with sync_playwright() as p:
        while True:
            n = random.randint(QA_TENANTS_PER_TEST_MIN, min(QA_TENANTS_PER_TEST_MAX, len(car_tenants)))
            sampled_tenants = random.sample(car_tenants, n)
            print(f"[qa-verifier] 這一輪抽 {len(sampled_tenants)} 家租戶: {', '.join(sampled_tenants)}")

            for i, tenant_id in enumerate(sampled_tenants):
                q = random.choice(questions)
                lang = random.choice(QA_TARGET_LANGS)
                print(f"[qa-verifier] 測試 tenant={tenant_id} | {q['id']} | lang={lang} | {q['question'][:20]}")

                mock_answer = q["expected_answer"] if QA_MOCK_GEMINI else None
                result = run_single_test(p, tenant_id, lang, q, mock_answer)
                append_result_excel(result)

                status = "成功" if result["success"] else "失敗"
                print(f"[qa-verifier]   -> {status}（耗時 {result['duration_ms']}ms）")

                # 同一輪內多個租戶共用同一把 Gemini Key 時，緊接著連續發送很容易撞到速率限制，
                # 所以租戶跟租戶之間錯開幾秒再測下一家（最後一家測完不用等）
                if i < len(sampled_tenants) - 1:
                    delay = random.uniform(QA_DELAY_BETWEEN_TESTS_MIN, QA_DELAY_BETWEEN_TESTS_MAX)
                    if wait_with_esc_check(delay, is_tty):
                        print("[qa-verifier] 偵測到 ESC，結束測試")
                        return

            # 每小時 2~4 輪 -> 平均間隔 15~30 分鐘，隨機化避免每次都固定間隔
            wait_seconds = 3600 / random.uniform(QA_TESTS_PER_HOUR_MIN, QA_TESTS_PER_HOUR_MAX)
            print(f"[qa-verifier] 休息約 {int(wait_seconds / 60)} 分鐘後繼續"
                  + ("（按 ESC 結束）..." if is_tty else "..."))

            if wait_with_esc_check(wait_seconds, is_tty):
                print("[qa-verifier] 偵測到 ESC，結束測試")
                return


if __name__ == "__main__":
    run_forever()
